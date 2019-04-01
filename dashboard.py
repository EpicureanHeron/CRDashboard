import openpyxl
import sys
from datetime import datetime
import time
import csv

start = time.time() #outside of everything because needs to start as soon as script starts?
now = datetime.now()

print('Beginning Dashboard Analysis...')

expenseAnalysis = openpyxl.load_workbook('expense_analysis.xlsx')
submitted = openpyxl.load_workbook('expense-submitted_reports.xlsx')
delegatesSetUp = openpyxl.load_workbook('Who_has_delegates_set_up.xlsx')
personReport = openpyxl.load_workbook('reference-person_report.xlsx')
approvalMethod = openpyxl.load_workbook('Beth_Approval_Method.xlsx')

sheet = expenseAnalysis.active
sheet2 = submitted.active
sheet3 = delegatesSetUp.active
sheet4 = personReport.active
sheet5 = approvalMethod.active

print('Spreadsheets are ready...')




def RRClist():
    
    pilotRRCs = ['ATHLX', 'AUXSV', 'AVPFN', 'CPPMX', 'FMXXX', 'OHRXX', 
    'OITXX', 'PSRXX', 'PUBSF', 'SUFIN', 'SVPFO', 'UHLSF', 'UMDXX', 'USERV']

    nonPilotRRCs = ['GPSTR', 'MNEXT', 'UMCXX', 'UMMXX', 'CCAPS', 'NURSG', 
    'OGCXX', 'UMRXX', 'PRESD', 'AUDIT', 'CSOMX', 'UEDUC', 'EQDIV', 'URELX',
    'AESXX', 'CEHDX', 'RSRCH', 'GRADX', 'AHCSH', 'AHSCI', 'HLSCI', 'CLAXX', 
    'CSENG', 'DESGN', 'LAWXX', 'LIBRX', 'STDAF', 'PUBHL', 'DENTX', 'HHHXX', 
    'PHARM', 'CFANS', 'AAPRV', 'MEDXX', 'VETMD', 'CBSXX', 'RGNTS']

    if len(sys.argv) > 1:
        includeList = []
        if sys.argv[1] == 'non-pilot':
            includeList = nonPilotRRCs
            return(includeList)
        else:
            for x in sys.argv[1:]:
                if x in pilotRRCs or x in nonPilotRRCs:

                    includeList.append(x)

            return(includeList)

    else:
        includeList = nonPilotRRCs + pilotRRCs
        return(includeList)

def approvalAnalysis(includeList):
    print("Analyzing Approvals...")
    i = includeList
    totalApprovers = 0
    totalEmail = 0
    totalSystem = 0
    uniqueApprovers = []
    RRCdata = {}
    
    
    for row in range(2, sheet5.max_row + 1):
       
        #print('passed')
        RRC = (sheet5['I' + str(row)].value)
        method = (sheet5['F' + str(row)].value)
        approverEmail = (sheet5['H' + str(row)].value) 

        if RRC in i:
           # print('%s %s %s' %(RRC, method, approverEmail))
            if method == 'MERC':
               # print('merc')
                totalSystem += 1
                RRCdata.setdefault(RRC + ' by system', 0)
                RRCdata[RRC + ' by system'] += 1 
                if approverEmail not in uniqueApprovers:
                    uniqueApprovers.append(approverEmail)
                    totalApprovers += 1

            elif method == 'EMAI':
           
                totalEmail += 1
                RRCdata.setdefault(RRC + ' by email', 0)
                RRCdata[RRC + ' by email'] += 1 
                if approverEmail not in uniqueApprovers:
                    uniqueApprovers.append(approverEmail)
                    totalApprovers += 1
            else: 
                if approverEmail not in uniqueApprovers:
                    uniqueApprovers.append(approverEmail)
                    totalApprovers += 1

    RRCdata.setdefault('By Email', totalEmail)
    RRCdata.setdefault('By System', totalSystem)
    RRCdata.setdefault('Unique approvers', totalApprovers)

    return(RRCdata)



def personAnalysis(includeList):
    print('Analyzing people data...')
    i = includeList
    totalPerson = 0
    RRCdata = {}

    for row in range(4, sheet4.max_row+1):

        RRC = (sheet4['AB' + str(row)].value)

        if RRC in i:
            RRCdata.setdefault(RRC, 0)
            RRCdata[RRC] += 1
            totalPerson += 1

    RRCdata.setdefault('Total Count', totalPerson)

    return(RRCdata)


def delegatesSetUpAnalysis(includeList):
    print('Analyzing expense owners with delegates set up...')
    i = includeList
    EOhasDelegate = 0
    RRCdata = {}

    for row in range(2, sheet3.max_row + 1):
        emailOfEO = (sheet3['E' + str(row)].value)
        RRC = (sheet3['C' + str(row)].value)

        if RRC in i:
            if emailOfEO != '':
                EOhasDelegate += 1
                RRCdata.setdefault(RRC, 0)
                RRCdata[RRC] += 1
               

    RRCdata.setdefault('Total EOs with Delegates', EOhasDelegate)
    return(RRCdata)


def submittedERsByDelegates(includeList):

    print("Calculating breakdown of ERs submitted by ER owner and by delegates...")

    i = includeList
    ERsByDelegates = 0
    ERsByExpenseOwners = 0
    RRCdata = {}

    for row in range(4, sheet2.max_row + 1):
        expenseOwner = (sheet2['E' + str(row)].value)
        expenseCreator = (sheet2['H' + str(row)].value)
        RRC = (sheet2['I' + str(row)].value)

        if RRC in i:
            if expenseCreator == expenseOwner:

                RRCdata.setdefault(RRC + ' by EO', 0)
                RRCdata[RRC + ' by EO'] += 1 
                ERsByExpenseOwners += 1
            else: 
                RRCdata.setdefault(RRC + ' by delegate', 0)
                RRCdata[RRC + ' by delegate'] += 1 
                ERsByDelegates += 1

    RRCdata.setdefault('Total by delegates', ERsByDelegates)
    RRCdata.setdefault('Total by expense owners', ERsByExpenseOwners)
    return(RRCdata)


def approvalTime(includeList):

    print('Calculating Approval Time...')

    i = includeList
    approvalTimes = []
    exportStatusList = ['Exported/Not Paid', 'Exported/Paid', 'Exported/Partially Paid']
    ERsAccountedFor = []
    lessThanThreeDaysCounter = 0


    for row in range(4, sheet.max_row + 1):


        ERID = (sheet['B' + str(row)].value)
        submittedDate = (sheet['C' + str(row)].value)
        exportedDate = (sheet['W' + str(row)].value)
        RRC = (sheet['Y' + str(row)].value)
        exportStatus = (sheet['V' + str(row)].value)


        if RRC in i:
            if exportStatus in exportStatusList:
                if  ERID not in ERsAccountedFor:
                    ERsAccountedFor.append(ERID)
                    try:
                        daysPassed = (exportedDate - submittedDate).days

                        approvalTimes.append(daysPassed)
    
                    except:
                        print('failed %s' % ERID)
                    if daysPassed <= 3:
                                lessThanThreeDaysCounter += 1

    average = sum(approvalTimes)/len(approvalTimes)

    lessThanThreePercent = lessThanThreeDaysCounter/len(approvalTimes)
    approvalTimeResults = {'Average': average,
                           'LessThan3Days': lessThanThreePercent}

    return(approvalTimeResults)

def spendAnalysis(includeList):

    print('Analyzing Spend...')

    i = includeList
            

    spendAnalysis = {'OOP': 0, 'Tcard': 0, "PD&M": 0}
    OOPalways = ['Per Diem Meals', 'Mileage']

    for row in range(4, sheet.max_row + 1):

        appAmount = (sheet['K' + str(row)].value)
        firmPaid = (sheet['X' + str(row)].value)
        RRC = (sheet['Y' + str(row)].value)
        appStatus = (sheet['U' + str(row)].value)
        expenseType = (sheet['G' + str(row)].value)

        if appStatus == 'Approved':
            if RRC in i:
            
                if firmPaid == 'Y':
                    spendAnalysis['Tcard'] += float(appAmount)
                else:
                    if expenseType in OOPalways:
                        spendAnalysis['PD&M'] += float(appAmount)
                    else:
                        spendAnalysis['OOP'] += float(appAmount)

    return(spendAnalysis)    
 
def ERsApprovedByRRC(includeList):

    print('Analyzing ERs by RRC...')

    i = includeList
    ERsAccountedFor = []
    RRCdata = {}
    TotalApproved = 0
  


    for row in range(4, sheet.max_row + 1):
        

        RRC = (sheet['Y' + str(row)].value)
        ERID = (sheet['B' + str(row)].value)
        appStatus = (sheet['U' + str(row)].value)

        if RRC in i:
        
            if appStatus == 'Approved':
                
                if ERID not in ERsAccountedFor:
                    TotalApproved += 1
                    ERsAccountedFor.append(ERID)

                    RRCdata.setdefault(RRC, 0)
                

                    RRCdata[RRC] += 1 
               

    RRCdata.setdefault('Total Approved', TotalApproved)

    return(RRCdata)

def ERsAffiliation(includeList):

    print('Analyzing Affiliation Data...')
    i = includeList
    ERsAccountedFor = []
    affiliationData = {}
    
    TotalSubmitted = 0
    for row in range(4, sheet.max_row + 1):
        ERID = (sheet['B' + str(row)].value)
        affl = (sheet['Z' + str(row)].value)
        RRC = (sheet['Y' + str(row)].value)
        if RRC in i:
            if ERID not in ERsAccountedFor:
                affiliationData.setdefault(affl, 0 )
                TotalSubmitted += 1
                ERsAccountedFor.append(ERID)
                affiliationData[affl] +=1

    affiliationData.setdefault('Total Submitted', TotalSubmitted)
    return(affiliationData)

def main():

    includeList = RRClist()

    a = ERsAffiliation(includeList)
    r = ERsApprovedByRRC(includeList)
    s = spendAnalysis(includeList)
    d = approvalTime(includeList)
    e = submittedERsByDelegates(includeList)
    d2 = delegatesSetUpAnalysis(includeList)
    p = personAnalysis(includeList)
    a2 = approvalAnalysis(includeList)

    name = str(datetime.now().date()) + ".csv"
    print('Preparing results...')
    with open(name, 'w') as f:  # Just use 'w' mode in 3.x
    
        w = csv.writer(f)
        w.writerow('ERS by Affilation')

        for row in a.items():
            w.writerow(row)
        
        w.writerow('------------')
        w.writerow('ERs approved by RRC')
        
        for row in r.items():
            w.writerow(row)

        w.writerow('------------')
        w.writerow('Spend Analysis')

        for row in s.items():
            w.writerow(row)
        
        w.writerow('------------')
        w.writerow('Approval Time')

        for row in d.items():
            w.writerow(row)
        
        w.writerow('------------')
        w.writerow('Submitted ERs')

        for row in e.items():
            w.writerow(row)
        
        w.writerow('------------')
        w.writerow('Delegates by Expense Owners')

        for row in d2.items():
            w.writerow(row)

        w.writerow('------------')
        w.writerow('Person Analysis')

        for row in p.items():
            w.writerow(row)

        w.writerow('------------')
        w.writerow('Approval Analysis')

        for row in a2.items():
            w.writerow(row)

    
    print('Done! Results in %s' %(name))

def test():
    includeList = RRClist()
    x = approvalAnalysis(includeList)
    print(x)

main()
# test()

log = open("log.txt", "a")
end = time.time()
totalTime = (end-start)
print('The script took %s seconds to run' % (totalTime))
log.write("date: " +str(now) + ", runtime: " + str(totalTime) + '\n')